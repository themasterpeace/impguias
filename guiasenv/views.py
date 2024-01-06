from datetime import date, timedelta
from pipes import Template
from urllib import request
from django.shortcuts import render, redirect
from django.views.generic import  ListView, CreateView, UpdateView, TemplateView
from django.contrib import messages
from django.urls import reverse_lazy

from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.messages.views import SuccessMessageMixin

from django.contrib.auth.decorators import login_required, permission_required
from django.views.decorators.csrf import csrf_exempt, csrf_protect
from django.utils.decorators import method_decorator
from openpyxl import Workbook
from openpyxl.styles import Alignment,Border,Font,PatternFill,Side
from django.contrib.auth import authenticate, login
from django.http.response import HttpResponse, JsonResponse

from django.db.models.functions import Coalesce
from django.db.models import Sum

from bases.views import SinPrivilegios
from .models import *
from .forms import *



class GuiaView( SinPrivilegios, ListView):
    permission_required = "guias.view_guia"
    model = GuiasEnv
    template_name = "guiasenv/guialist.html"
    context_object_name = "obj"
    login_url = "bases:login"


    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)
    
    def post(self, request, *args, **kwargs):
        data = []
        try:
            action = request.POST['action']
            if action == 'searchdata':
                data = []
                for i in GuiasEnv.objects.all():
                    data.append(i.toJSON())
            else:
                data['error'] = 'Ha ocurrido un error'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)
        

class GuiaNew(SuccessMessageMixin, SinPrivilegios,\
    CreateView):
    permission_required = "guiasenv.add_guiasenv"
    model = GuiasEnv
    template_name = "guiasenv/guianew.html"
    context_object_name = "obj"
    form_class = GuiaForm
    succes_url = reverse_lazy("guiasenv:guialist")
    success_message = 'CORRELATIVO DE GUIAS IMPRESAS AGREGADO CORRECTAMENTE'
    login_url = "bases:login"

    def form_valid(self, form):
        form.instance.uc = self.request.user
        return super().form_valid(form)


class GuiaEdit(SuccessMessageMixin, SinPrivilegios, UpdateView):
    permission_required = "guiasenv.update_guia"
    model = GuiasEnv
    template_name="guiasenv/guianew.html"
    context_object_name="obj"
    form_class = GuiaForm
    success_url=reverse_lazy("guiasenv:guialist")
    success_message = 'CORRELATIVO DE GUIAS IMPRESAR ACTUALIZADO CORRECTAMENTE'
    login_url = "bases:login"

    def form_valid(self, form):
        form.instance.um = self.request.user.id
        return super().form_valid(form)


@login_required(login_url="/login/")
@permission_required("guaisenv.change_guiasenv", login_url="/login/")
def entregado(request, id):
    codigo =GuiasEnv.objects.filter(pk=id).first()
    
    if request.method == 'POST':
        if codigo:
            codigo.entregado = not codigo.entregado
            codigo.save()
            return HttpResponse("OK")
        return HttpResponse("FAIL")

    return HttpResponse("FAIL")

    return render(request, template_name, contexto)

"""Seccion de Reportes de impresion de guias """

class Reportes(LoginRequiredMixin, TemplateView):
    template_name = 'guiasenv/reportes.html'
    login_url = 'bases:login'

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)
    
    def post(self, request, *args, **kwargs):
        data = []
        try:
            action = request.POST['action']
            if action == 'search_report':
                data = []
                start_date = request.POST.get('start_date', '')
                end_date = request.POST.get('end_date', '')
                search = GuiasEnv.objects.all()
                if len(start_date) and len(end_date):
                    search = search.filter(fecha__range=[start_date, end_date])
                for s in search:
                    data.append([
                        s.id,
                        s.fecha,
                        s.codigo,
                        s.cliente,
                        s.tipo_envio,
                        s.numini,
                        s.numfin,
                        format(s.totenvio, '.2f'),
                    ])
                
                total = search.aggregate(r=Coalesce(Sum('totenvio'), 0)).get('r')

                data.append([
                    '---',
                    '---',
                    '---',
                    '---',
                    '---',
                    '---',
                    '---',
                    format(total, '.2f'),
                ])

            else:
                data['error'] = 'Ha ocurrido un error'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)
        

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = ReportForm()

        return context

class ReporteGeneralExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        campo = request.GET.get('campo')
        clientes = GuiasEnv.objects.all()
        
        wb = Workbook()
        ws = wb.active

        #Crear el título en la hoja
        ws['A1'].alignment = Alignment(horizontal = "center",vertical = "center")
        ws['A1'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") ) 
        ws['A1'].fill = PatternFill(start_color = '66FFCC', end_color = '66FFCC', fill_type = "solid")
        ws['A1'].font = Font(name = 'Calibri', size = 12, bold = True)
        ws['A1'] = 'REPORTE GUIAS RUTAS Y AGENCIAS'

        ws.merge_cells('A1:I1')
        ws.row_dimensions[1].height = 25

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 20


        ws['A3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['A3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['A3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['A3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['A3'] = 'FECHA'

        ws['B3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['B3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['B3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['B3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['B3'] = 'CODIGO'
        
        ws['C3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['C3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['C3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['C3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['C3'] = 'CLIENTE'

        ws['D3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['D3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                   top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['D3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['D3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['D3'] = 'TIPO ENVIO'

        ws['E3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['E3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['E3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['E3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['E3'] = 'NUMERO INICIAL'

        ws['F3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['F3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['F3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['F3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['F3'] = 'NUMERO FINAL'

        ws['G3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['G3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['G3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['G3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['G3'] = 'TOTAL ENVIADO'

        ws['H3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['H3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['H3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['H3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['H3'] = 'FORMA PAGO'

        ws['I3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['I3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['I3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['I3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['I3'] = 'ENTREGADO'

        cont = 4 

        for cliente in clientes:
            ws.cell(row = cont, column=1).value = cliente.fecha
            ws.cell(row = cont, column=2).value = cliente.codigo
            ws.cell(row = cont, column=3).value = cliente.cliente
            ws.cell(row = cont, column=4).value = cliente.tipo_envio
            ws.cell(row = cont, column=5).value = cliente.numini
            ws.cell(row = cont, column=6).value = cliente.numfin
            ws.cell(row = cont, column=7).value = cliente.totenvio
            ws.cell(row = cont, column=8).value = cliente.fpago
            ws.cell(row = cont, column=9).value = cliente.entregado
            cont+=1

        nombre_archivo = "GuiasImpresas.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition']= content
        wb.save(response)
        return response

class ReporteClienteExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        campo = request.GET.get('campo')
        clientes = GuiasEnv.objects.filter(codigo = campo)
        
        wb = Workbook()
        ws = wb.active

        #Crear el título en la hoja
        ws['A1'].alignment = Alignment(horizontal = "center",vertical = "center")
        ws['A1'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") ) 
        ws['A1'].fill = PatternFill(start_color = '66FFCC', end_color = '66FFCC', fill_type = "solid")
        ws['A1'].font = Font(name = 'Calibri', size = 12, bold = True)
        ws['A1'] = 'REPORTE GUIAS IMPRESAS POR CLIENTE'

        ws.merge_cells('A1:I1')
        ws.row_dimensions[1].height = 25

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 20


        ws['A3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['A3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['A3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['A3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['A3'] = 'FECHA'

        ws['B3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['B3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['B3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['B3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['B3'] = 'CODIGO'
        
        ws['C3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['C3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['C3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['C3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['C3'] = 'CLIENTE'

        ws['D3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['D3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                   top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['D3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['D3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['D3'] = 'TIPO ENVIO'

        ws['E3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['E3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['E3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['E3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['E3'] = 'NUMERO INICIAL'

        ws['F3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['F3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['F3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['F3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['F3'] = 'NUMERO FINAL'

        ws['G3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['G3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['G3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['G3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['G3'] = 'TOTAL ENVIADO'

        ws['H3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['H3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['H3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['H3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['H3'] = 'FORMA PAGO'

        ws['I3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['I3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['I3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['I3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['I3'] = 'ENTREGADO'

        cont = 4 

        for cliente in clientes:
            ws.cell(row = cont, column=1).value = cliente.fecha
            ws.cell(row = cont, column=2).value = cliente.codigo
            ws.cell(row = cont, column=3).value = cliente.cliente
            ws.cell(row = cont, column=4).value = cliente.tipo_envio
            ws.cell(row = cont, column=5).value = cliente.numini
            ws.cell(row = cont, column=6).value = cliente.numfin
            ws.cell(row = cont, column=7).value = cliente.totenvio
            ws.cell(row = cont, column=8).value = cliente.fpago
            ws.cell(row = cont, column=9).value = cliente.entregado
            cont+=1

        nombre_archivo = "GuiasImpresas.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition']= content
        wb.save(response)
        return response

class ReporteFechaExcel(TemplateView):

    def get(self, request, *args, **kwargs):
        #campo = request.GET.get('campo')
        fecha_desde = request.GET.get('fecha')
        fecha_hasta = request.GET.get('fecha_hasta')
        #clientes= GuiasEnv.objects.all()
        clientes = GuiasEnv.objects.filter(fecha__range=[fecha_desde, fecha_hasta])
               
        wb = Workbook()
        ws = wb.active

        #Crear el título en la hoja
        ws['A1'].alignment = Alignment(horizontal = "center",vertical = "center")
        ws['A1'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") ) 
        ws['A1'].fill = PatternFill(start_color = '66FFCC', end_color = '66FFCC', fill_type = "solid")
        ws['A1'].font = Font(name = 'Calibri', size = 12, bold = True)
        ws['A1'] = 'REPORTE GUIAS IMPRESAS POR FECHA'

        ws.merge_cells('A1:I1')
        ws.row_dimensions[1].height = 25

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 20


        ws['A3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['A3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['A3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['A3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['A3'] = 'FECHA'

        ws['B3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['B3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['B3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['B3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['B3'] = 'CODIGO'
        
        ws['C3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['C3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['C3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['C3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['C3'] = 'CLIENTE'

        ws['D3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['D3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                   top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['D3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['D3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['D3'] = 'TIPO ENVIO'

        ws['E3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['E3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['E3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['E3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['E3'] = 'NUMERO INICIAL'

        ws['F3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['F3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['F3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['F3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['F3'] = 'NUMERO FINAL'

        ws['G3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['G3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['G3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['G3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['G3'] = 'TOTAL ENVIADO'

        ws['H3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['H3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['H3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['H3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['H3'] = 'FORMA PAGO'

        ws['I3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['I3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['I3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['I3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['I3'] = 'ENTREGADO'

        cont = 4 

        for cliente in clientes:
            ws.cell(row = cont, column=1).value = cliente.fecha
            ws.cell(row = cont, column=2).value = cliente.codigo
            ws.cell(row = cont, column=3).value = cliente.cliente
            ws.cell(row = cont, column=4).value = cliente.tipo_envio
            ws.cell(row = cont, column=5).value = cliente.numini
            ws.cell(row = cont, column=6).value = cliente.numfin
            ws.cell(row = cont, column=7).value = cliente.totenvio
            ws.cell(row = cont, column=8).value = cliente.fpago
            ws.cell(row = cont, column=9).value = cliente.entregado
            cont+=1

        nombre_archivo = "GuiasImpresas.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition']= content
        wb.save(response)
        return response



"""Zona de operaciones CRUD de listado de clientes """
class ClienteListView(SinPrivilegios, ListView):
    permission_required = "guias.view_cliente"
    model = Cliente
    template_name = "guiasenv/clientelist.html"
    context_object_name = "obj"
    login_url = "bases:login"


    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)
    
    def post(self, request, *args, **kwargs):

        data = []
        try:
            action = request.POST['action']
            if action == 'searchdata':
                data = []
                for i in Cliente.objects.all():
                    data.append(i.toJSON())
            else:
                data['error'] = 'Ha ocurrido un error'
        except Exception as e:
            print(data)
            data['error'] = str(e)
        return JsonResponse(data, safe=False)

class ClienteNew(SuccessMessageMixin, SinPrivilegios, CreateView):
    
    permission_required = "guiasenv.create_cliente"
    model = Cliente
    template_name = "guiasenv/clientenew.html"
    context_object_name = "obj"
    form_class = ClienteForm
    succes_url = reverse_lazy("guiasenv:clientelist")
    success_message = 'CLIENTE REGISTRADO EXITOSAMENTE'
    login_url = "bases:login"

    def form_valid(self, form):
        form.instance.uc = self.request.user
        return super().form_valid(form)

class ClienteEdit(SuccessMessageMixin, SinPrivilegios, UpdateView):
    permission_required = "guiasenv.update_cliente"
    model = Cliente
    template_name="guiasenv/clientenew.html"
    context_object_name="obj"
    form_class = ClienteForm
    success_url=reverse_lazy("guiasenv:clientelist")
    success_message = 'CLIENTE ACTUALIZADO EXITOSAMENTE'
    login_url = "bases:login"

    def form_valid(self, form):
        user_instance=self.request.user
        form.instance.um = user_instance
        return super().form_valid(form)

def registro(request):
    data = {
        'form': CustomUserCreationForm()
    }

    if request.method == 'POST':
        formulario = CustomUserCreationForm(data=request.POST)
        if formulario.is_valid():
            formulario.save()
            user = authenticate(
                username=formulario.cleaned_data["username"], password=formulario.cleaned_data["password1"])
            login(request, user)
            messages.success(request, "Te has registrado exitosamente")
            #redirigir al home
            return redirect(to="base:home")
        data["form"] = formulario
    return render(request, 'registration/registro.html', data)

#Enviar a excel informacion de tabla dbf

import pandas as pd
from django.http import HttpResponse
from django.views import View
from dbfread import DBF

class ReporteHistorico(View):
    def get(self, request, *args, **kwargs):
        fecha_inicial = request.GET.get('fecha_inicial')
        fecha_final = request.GET.get('fecha_final')

        # Ruta al archivo DBF
        ruta_dbf = 'C:/Users/mesquite/Desktop/dumpsisam/envioshistorico.dbf'

        # Lee el archivo DBF y convierte a DataFrame
        dbf_table = DBF(ruta_dbf)
        df = pd.DataFrame(iter(dbf_table))

        # Filtra los datos por rango de fechas
        df['fecha'] = pd.to_datetime(df['fecha'])
        df = df[(df['fecha'] >= fecha_inicial) & (df['fecha'] <= fecha_final)]

        # Escribir DataFrame en un archivo Excel
        excel_output_path = 'ruta_del_archivo.xlsx'
        df.to_excel(excel_output_path, index=False, engine='openpyxl')

        # Devolver el archivo Excel como respuesta
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=reporte.xlsx'
        df.to_excel(response, index=False, engine='openpyxl')

        return response

