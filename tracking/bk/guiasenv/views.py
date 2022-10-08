import time
from pipes import Template
from urllib import request
from django.shortcuts import render, redirect
from django.views.generic import  ListView, CreateView, UpdateView, TemplateView
from django.contrib import messages
from django.urls import reverse_lazy

from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.messages.views import SuccessMessageMixin

from django.contrib.auth.decorators import login_required, permission_required
from django.views.decorators.csrf import csrf_exempt, csrf_protect
from django.utils.decorators import method_decorator
from openpyxl import Workbook
from django.http.response import HttpResponse, JsonResponse

from bases.views import SinPrivilegios
from .models import *
from .forms import GuiaForm



class GuiaView( SinPrivilegios, ListView):
    permission_required = "guias.view_guia"
    model = GuiasEnv
    template_name = "guiasenv/guialist.html"
    context_object_name = "obj"
    login_url = "bases:login"


    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)
    
    def post(self, request, *args, **kwargs):
        data = []
        try:
            action = request.POST['action']
            if action == 'searchdata':
                data = []
                for i in GuiasEnv.objects.all():
                    data.append(i.toJSON())
            else:
                data['error'] = 'Ha ocurrido un error'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)
        



class GuiaNew(SuccessMessageMixin, SinPrivilegios,\
    CreateView):
    permission_required = "guiasenv.create_guia"
    model = GuiasEnv
    template_name = "guiasenv/guianew.html"
    context_object_name = "obj"
    form_class = GuiaForm
    succes_url = reverse_lazy("guiasenv:guialist")
    success_message = 'CORRELATIVO DE GUIAS IMPRESAS AGREGADO CORRECTAMENTE'
    login_url = "bases:login"

    def form_valid(self, form):
        form.instance.uc = self.request.user
        return super().form_valid(form)


class GuiaEdit(SuccessMessageMixin, SinPrivilegios, UpdateView):
    permission_required = "guiasenv.update_guia"
    model = GuiasEnv
    template_name="guiasenv/guianew.html"
    context_object_name="obj"
    form_class = GuiaForm
    success_url=reverse_lazy("guiasenv:guialist")
    success_message = 'CORRELATIVO DE GUIAS IMPRESAR ACTUALIZADO CORRECTAMENTE'
    login_url = "bases:login"

    def form_valid(self, form):
        form.instance.um = self.request.user.id
        return super().form_valid(form)


@login_required(login_url="/login/")
@permission_required("guaisenv.change_guiasenv", login_url="/login/")
def entregado(request, id):
    codigo =GuiasEnv.objects.filter(pk=id).first()
    
    if request.method == 'POST':
        if codigo:
            codigo.entregado = not codigo.entregado
            codigo.save()
            return HttpResponse("OK")
        return HttpResponse("FAIL")

    return HttpResponse("FAIL")

    return render(request, template_name, contexto)



class ReporteClienteExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        clientes = GuiasEnv.objects.all()
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE DE CLIENTES'

        ws.merge_cells('B1:E1')

        ws['B3'] = 'FECHA'
        ws['C3'] = 'CODIGO'
        ws['D3'] = 'CLIENTE'
        ws['E3'] = 'TIPO ENVIO'
        ws['F3'] = 'NUMERO INICIAL'
        ws['G3'] = 'NUMERO FINAL'
        ws['H3'] = 'TOTAL ENVIADO'
        ws['I3'] = 'FORMA PAGO'
        ws['J3'] = 'ENTREGADO'

        cont = 9 

        for cliente in clientes:
            ws.cell(row = cont, column=2).value = cliente.fecha
            ws.cell(row = cont, column=3).value = cliente.codigo
            ws.cell(row = cont, column=4).value = cliente.cliente
            ws.cell(row = cont, column=5).value = cliente.tipo_envio
            ws.cell(row = cont, column=6).value = cliente.numini
            ws.cell(row = cont, column=7).value = cliente.numfin
            ws.cell(row = cont, column=8).value = cliente.totenvio
            ws.cell(row = cont, column=9).value = cliente.fpago
            ws.cell(row = cont, column=10).value = cliente.entregado
            cont+=1

        nombre_archivo = "GuiasImpresas.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition']= content
        wb.save(response)
        return response
