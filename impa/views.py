import time

from re import  template
from django.http import request 
from django.shortcuts import redirect, render, get_object_or_404
from .models import *
from .forms import * 
from django.urls import reverse_lazy
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.messages.views import SuccessMessageMixin
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, TemplateView
from django.contrib import messages 
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required, permission_required
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt, csrf_protect
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.mixins import LoginRequiredMixin
from openpyxl import Workbook
import calendar
from calendar import month_name
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from guiasenv.views import ClienteListView

from bases.views import SinPrivilegios, BaseCreate

# Create your views here.

class impview(LoginRequiredMixin, ListView):
    model = ImpGuias
    template_name ="impa/listimp.html"
    context_object_name = "obj"
    login_url = "bases:login"

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)
    
    def post(self, request, *args, **kwargs):
        data = []
        try:
            action = request.POST['action']
            if action == 'searchdata':
                data = []
                for i in ImpGuias.objects.all():
                    data.append(i.toJSON())
            else:
                data['error'] = 'Ha ocurrido un error'
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data, safe=False)


class impnew(BaseCreate):
    permission_required="impa.create_impguias"
    model = ImpGuias
    template_name = "impa/newimp.html"
    form_class = ImpresionForm
    success_url = reverse_lazy("impa:listimp")
    success_message = "IMPRESION DE GUIAS MADRE ENVIADA EXITOSAMENTE"
    
        
class impedit(LoginRequiredMixin, UpdateView):
    model = ImpGuias
    template_name="impa/newimp.html"
    context_object_name="obj"
    form_class = ImpresionForm
    success_url=reverse_lazy("imp:listimp")
    login_url = "bases:login"

    def form_valid(self, form):
        form.instance.um = self.request.user.id
        return super().form_valid(form)

class impdel(LoginRequiredMixin, DeleteView):
    model = ImpGuias
    template_name='impa/eliminar.html'
    context_object_name='obj'
    success_url=reverse_lazy("impa:listimp")
    login_url = "bases:login"

class buscarcli(ClienteListView):
    template_name='link/buscar_cli.html'

class buscardes(ClienteListView):
    template_name='link/buscar_des.html'

#class buscarprod(tarifarioview):
#    template_name='link/buscar_prod.html'
    
class ReportImp(TemplateView):
    def get(self, request, *args, **kwargsa):
        campo = request.GET.get('campo')
        clientes = ImpGuias.objects.all()

        wb = Workbook()
        ws = wb.active

        #Seccion de titulo de la hoja electronica
        ws['A1'].alignment = Alignment(horizontal = "center",vertical = "center")
        ws['A1'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") ) 
        ws['A1'].fill = PatternFill(start_color = '66FFCC', end_color = '66FFCC', fill_type = "solid")
        ws['A1'].font = Font(name = 'Calibri', size = 12, bold = True)
        ws['A1'] = 'REPORTE GUIAS IMPRESAS A CLIENTES'

        ws.merge_cells('A1:I1')
        ws.row_dimensions[1].height = 25

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20

        ws['A3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['A3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['A3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['A3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['A3'] = 'FECHA'

        ws['B3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['B3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['B3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['B3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['B3'] = 'CODIGO'
        
        ws['C3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['C3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['C3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['C3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['C3'] = 'CLIENTE'

        ws['D3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['D3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['D3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['D3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['D3'] = 'NUMERO INICIAL'

        ws['E3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['E3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['E3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['E3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['E3'] = 'NUMERO FINAL'

        ws['F3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['F3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['F3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['F3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['F3'] = 'FORMA PAGO'
        

        ws['G3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['G3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['G3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['G3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['G3'] = 'TOTAL IMPRESO'

        cont = 4 

        for cliente in clientes:
            ws.cell(row = cont, column=1).value = cliente.fecha
            ws.cell(row = cont, column=2).value = cliente.codigo_cliente
            ws.cell(row = cont, column=3).value = cliente.remitente
            ws.cell(row = cont, column=4).value = cliente.numini
            ws.cell(row = cont, column=5).value = cliente.numfin
            ws.cell(row = cont, column=6).value = cliente.fpago
            ws.cell(row = cont, column=7).value = cliente.totalimp
            cont+=1

        nombre_archivo = "GuiasImpresasclientes.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition']= content
        wb.save(response)
        return response
    
class ReportClie(TemplateView):
    def get(self, request, *args, **kwargsa):
        campo = request.GET.get('campo')
        clientes = ImpGuias.objects.filter(codigo_cliente=campo)

        wb = Workbook()
        ws = wb.active

        #Seccion de titulo de la hoja electronica
        ws['A1'].alignment = Alignment(horizontal = "center",vertical = "center")
        ws['A1'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") ) 
        ws['A1'].fill = PatternFill(start_color = '66FFCC', end_color = '66FFCC', fill_type = "solid")
        ws['A1'].font = Font(name = 'Calibri', size = 12, bold = True)
        ws['A1'] = 'REPORTE GUIAS IMPRESAS A CLIENTES'

        ws.merge_cells('A1:I1')
        ws.row_dimensions[1].height = 25

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20

        ws['A3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['A3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['A3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['A3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['A3'] = 'FECHA'

        ws['B3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['B3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['B3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['B3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['B3'] = 'CODIGO'
        
        ws['C3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['C3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['C3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['C3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['C3'] = 'CLIENTE'

        ws['D3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['D3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['D3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['D3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['D3'] = 'NUMERO INICIAL'

        ws['E3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['E3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['E3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['E3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['E3'] = 'NUMERO FINAL'

        ws['F3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['F3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['F3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['F3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['F3'] = 'FORMA PAGO'
        

        ws['G3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['G3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['G3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['G3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['G3'] = 'TOTAL IMPRESO'

        cont = 4 

        for cliente in clientes:
            ws.cell(row = cont, column=1).value = cliente.fecha
            ws.cell(row = cont, column=2).value = cliente.codigo_cliente
            ws.cell(row = cont, column=3).value = cliente.remitente
            ws.cell(row = cont, column=4).value = cliente.numini
            ws.cell(row = cont, column=5).value = cliente.numfin
            ws.cell(row = cont, column=6).value = cliente.fpago
            ws.cell(row = cont, column=7).value = cliente.totalimp
            cont+=1

        nombre_archivo = f"cliente_{campo}.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition']= content
        wb.save(response)
        return response
    
class ReportDate(TemplateView):
    def get(self, request, *args, **kwargsa):
        fecha_desde = request.GET.get('fecha')
        fecha_hasta = request.GET.get('fecha_hasta')
        clientes = ImpGuias.objects.filter(fecha__range=[fecha_desde, fecha_hasta])

        wb = Workbook()
        ws = wb.active

        #Seccion de titulo de la hoja electronica
        ws['A1'].alignment = Alignment(horizontal = "center",vertical = "center")
        ws['A1'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") ) 
        ws['A1'].fill = PatternFill(start_color = '66FFCC', end_color = '66FFCC', fill_type = "solid")
        ws['A1'].font = Font(name = 'Calibri', size = 12, bold = True)
        ws['A1'] = 'REPORTE GUIAS IMPRESAS A CLIENTES'

        ws.merge_cells('A1:I1')
        ws.row_dimensions[1].height = 25

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20

        ws['A3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['A3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['A3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['A3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['A3'] = 'MES'

        ws['B3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['B3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['B3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['B3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['B3'] = 'CODIGO'
        
        ws['C3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['C3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['C3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['C3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['C3'] = 'CLIENTE'

        ws['D3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['D3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['D3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['D3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['D3'] = 'NUMERO INICIAL'

        ws['E3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['E3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['E3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['E3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['E3'] = 'NUMERO FINAL'

        ws['F3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['F3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['F3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['F3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['F3'] = 'FORMA PAGO'
        

        ws['G3'].alignment = Alignment(horizontal = "center", vertical = "center")
        ws['G3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
        ws['G3'].fill = PatternFill(start_color = '66CFCC', end_color = '66CFCC', fill_type = "solid")
        ws['G3'].font = Font(name = 'Calibro', size = 10, bold = True)
        ws['G3'] = 'TOTAL IMPRESO'

        cont = 4 

        for cliente in clientes:
            nombre_mes = month_name[cliente.fecha.month]
            ws.cell(row = cont, column=1).value = cliente.fecha
            ws.cell(row = cont, column=2).value = cliente.codigo_cliente
            ws.cell(row = cont, column=3).value = cliente.remitente
            ws.cell(row = cont, column=4).value = cliente.numini
            ws.cell(row = cont, column=5).value = cliente.numfin
            ws.cell(row = cont, column=6).value = cliente.fpago
            ws.cell(row = cont, column=7).value = cliente.totalimp
            cont+=1

        nombre_mes = calendar.month_name[int(fecha_desde.split('-')[1])]
        nombre_archivo = f"Mes_Generado_{nombre_mes}.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition']= content
        wb.save(response)
        return response